"""
Stage-based scrap costing: instead of one flat cost per model, this looks
up the cost of the SPECIFIC build stage a unit had reached when it was
scrapped - a unit that fails early (e.g. at an SMT station) costs far less
than one that fails late (e.g. at Final VI), since fewer components/labor
have gone into it.

Data comes from "Final Cost for Dashboard.xlsx":
- Per-product-code sheets (EP1110, EP1140, ...): Stage -> already-rolled-up
  price for that stage (you did the BOM+inventory rollup yourself)
- "Micro Inverters" / "Batteries" sheets: Model -> Product Code mapping
- "Summary" sheet: flat FG cost per model, used as a fallback when a
  specific stage price can't be found

Station -> Stage classification is NOT in the spreadsheet's cell data (it's
two embedded images), so it's hardcoded below exactly as given.
"""
import collections
import openpyxl


# ---- Station -> Stage classification (from the embedded images in the
# cost file's Sheet1) - hardcoded since this isn't in the actual cell data.
STATION_TO_STAGE = {
    "Micro Inverters": {
        "SMT": ["PCB LOADING", "SCREEN PRINTER", "SOLDER PASTE INSPECTION", "POST AOI", "PCB DEPANEL", "AOI", "SPI"],
        "MI": ["FIXTURE PCB LINKING", "API MI", "API", "PVI"],
        "PCBA": ["PCB VI", "PCB_VI", "LEAD HEIGHT MEASUREMENT", "ICT", "FT", "WAVE SOLDER", "PW-VI"],
        "M-Assembly": ["ENCLOSURE LASER MARKING", "GLUE HEIGHT MEASUREMENT", "CAROUSEL TRACEBILITY", "CAROUSEL TRACEABILITY",
                       "POTTING", "HT", "GT", "WAVE 3 FIXTURE LINKING", "CAP USW", "FINAL VI", "PVI"],
        # "Wave 3" stations (per explicit confirmation) - CC VI / CVI are
        # the same as Carousel VI, just abbreviated differently.
        "FG": ["ENROUTE CHECK", "OBA", "PACK 2 BOX", "CC VI", "CAROUSEL VI", "CVI",
               "ACF LABEL PRINTING", "CAN LABEL PRINTING", "SN3 GEN & LINKING",
               "PACKING", "WAVE 3 TRACEABILITY", "PCU DE-LINK",
               "BMCC LABEL PRINTING", "CC BOTTOM", "MANUAL MI ASSEMBLY", "AUTO WELDING",
               "FIXTURE LINKING MANUAL", "ASSEMBLE BASE & LASER WELDING",
               "POTTING & UNIT MANAUL LOADING", "PACK 2 PALLET",
               "PCBA HEAT STACKING LID", "SN3 LABEL PRINT", "FVI",
               "GLUE HEIGHT MEASUREMENT & EMC STATION", "GLUE HEIGHT MEASUREMENT & EMC", "SN LINKING"],
    },
    "Batteries": {
        # Some battery products (e.g. EP2100) DO go through SMT/MI/PCBA
        # stages before final assembly - these were missing entirely
        # before, which meant those scrap events always fell back to the
        # much more expensive flat FG cost instead of their real (cheaper)
        # stage cost. Using the same station names as Micro Inverters since
        # these are shared physical stations/machines.
        "SMT": ["PCB LOADING", "SCREEN PRINTER", "SOLDER PASTE INSPECTION", "POST AOI", "PCB DEPANEL", "AOI", "SPI"],
        "MI": ["FIXTURE PCB LINKING", "API MI", "API", "PVI"],
        "PCBA": ["PCB VI", "PCB_VI", "LEAD HEIGHT MEASUREMENT", "ICT", "FT", "WAVE SOLDER", "PW-VI"],
        "Assembly Phantom": ["UNIT SN GENERATION", "GASKET VI", "SN3 LINKING", "ASSEMBLY MOUNT BOLT",
                              "INSTALL WIRING COMPARTMENT VI", "CABLE CONNECTIONS", "CABLE CONNECTIONS VI",
                              "BATTERY PACK LINKING", "PCBA LINKING"],
        "M-Assembly": ["HT", "PCU CONNECTION", "PCU CONNECT TRACEBILITY", "PCU CONNECT TRACEABILITY", "HEATSINK VI",
                       "HEAT SINK VI TRACEBILITY", "HEAT SINK VI TRACEABILITY", "BST", "CCD TOP,SIDE", "UNIT LABEL",
                       "OPEN TERMINAL BULKHEAD", "ACFDF SCREW", "DEAD FRONT ACF",
                       "INSTALLING WIRING COMPARTMENT", "FINAL VI"],
        # "Wave 3" stations (per explicit confirmation) - CC VI / CVI are
        # the same as Carousel VI, just abbreviated differently.
        "FG": ["ENROUTE CHECK", "OBA", "UNIT PACKING TRACEBILITY", "UNIT PACKING TRACEABILITY", "ACCESSORY BAG",
               "PACK 2 BOX", "CC VI", "CAROUSEL VI", "CVI", "ACF LABEL PRINTING", "CAN LABEL PRINTING",
               "SN3 GEN & LINKING", "PACKING", "WAVE 3 TRACEABILITY", "PCU DE-LINK",
               "BMCC LABEL PRINTING", "CC BOTTOM", "MANUAL MI ASSEMBLY", "AUTO WELDING",
               "FIXTURE LINKING MANUAL", "ASSEMBLE BASE & LASER WELDING",
               "POTTING & UNIT MANAUL LOADING", "PACK 2 PALLET",
               "PCBA HEAT STACKING LID", "SN3 LABEL PRINT", "FVI",
               "GLUE HEIGHT MEASUREMENT & EMC STATION", "GLUE HEIGHT MEASUREMENT & EMC", "SN LINKING"],
    },
}

# Reverse lookup: station name (normalized) -> stage, per family - built once.
_STATION_LOOKUP = {}
for family, stages in STATION_TO_STAGE.items():
    _STATION_LOOKUP[family] = {}
    for stage, stations in stages.items():
        for s in stations:
            key = s.strip().upper()
            # if a station appears in multiple stages (e.g. PVI), first
            # occurrence wins - same "don't silently overwrite" principle
            # used elsewhere in this codebase
            if key not in _STATION_LOOKUP[family]:
                _STATION_LOOKUP[family][key] = stage


def get_stage_for_station(family, station):
    if not family or not station:
        return None
    return _STATION_LOOKUP.get(family, {}).get(str(station).strip().upper())


def load_stage_cost_data(filepath):
    """
    Returns a dict with everything needed for per-row stage costing:
      model_to_product_code: {normalized model text: product code}
      product_code_to_family: {product code: 'Micro Inverters' | 'Batteries'}
      product_code_to_stage_prices: {product code: {stage name: price}}
      product_code_to_flat_cost: {product code: flat FG cost} (fallback)
    Returns None if the file doesn't exist yet.
    """
    import os
    if not filepath or not os.path.exists(filepath):
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)

    model_to_product_code = {}
    product_code_to_family = {}
    for family in ("Micro Inverters", "Batteries"):
        if family not in wb.sheetnames:
            continue
        ws = wb[family]
        for row in ws.iter_rows(min_row=4, values_only=True):
            product_code, model = row[0], row[1]
            if not product_code or not model:
                continue
            product_code = str(product_code).strip()
            model = str(model).strip()
            if product_code == "Product Code" or model == "Model":
                continue  # skip a stray repeated header row
            if product_code not in ("Micro Inverters", "Batteries") and model:
                if model not in model_to_product_code:
                    model_to_product_code[model] = product_code
                if product_code not in product_code_to_family:
                    product_code_to_family[product_code] = family

    product_code_to_flat_cost = {}
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            product_code, model, total = row[0], row[1], row[2]
            if product_code and total is not None:
                product_code_to_flat_cost[str(product_code).strip()] = float(total)

    product_code_to_stage_prices = {}
    skip_sheets = {"Summary", "Micro Inverters", "Batteries", "Sheet1", "Sheet2", "Sheet3"}
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        stages = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            stage, phantom_code, desc, price = row[0], row[1], row[2], row[3]
            if stage and price is not None:
                stages[str(stage).strip()] = float(price)
        if stages:
            product_code_to_stage_prices[sheet_name.strip()] = stages

    return {
        "model_to_product_code": model_to_product_code,
        "product_code_to_family": product_code_to_family,
        "product_code_to_stage_prices": product_code_to_stage_prices,
        "product_code_to_flat_cost": product_code_to_flat_cost,
    }


def get_product_code(model_text, data):
    """Matches a defect log MODEL string to a Product Code, trying comma
    segments the same way the flat-cost matcher does (handles both
    'E/M Assembly, IQ8BL-DOM' and 'IQ9N-A-INT, MADE IN US...' patterns)."""
    if not model_text or not data:
        return None
    model_text = str(model_text).strip()
    lookup = data["model_to_product_code"]
    if model_text in lookup:
        return lookup[model_text]
    if "," in model_text:
        for segment in model_text.split(","):
            segment = segment.strip()
            if segment in lookup:
                return lookup[segment]
    return None


def get_scrap_cost(model_text, station, data):
    """
    Returns (cost, detail) for one scrap event, where detail explains how
    the cost was determined - useful for debugging/verifying.
    Falls back to the flat FG cost if the specific stage can't be
    determined (unknown station, or that stage isn't priced for this model).
    Returns (None, reason) if nothing at all could be matched.
    """
    product_code = get_product_code(model_text, data)
    if product_code is None:
        return None, "model not matched to a product code"

    family = data["product_code_to_family"].get(product_code)
    stage = get_stage_for_station(family, station)

    if stage:
        stage_prices = data["product_code_to_stage_prices"].get(product_code, {})
        if stage in stage_prices:
            return stage_prices[stage], f"{product_code} / {stage} stage"

    flat_cost = data["product_code_to_flat_cost"].get(product_code)
    if flat_cost is not None:
        reason = "flat FG cost (stage not found)" if stage else "flat FG cost (station not classified)"
        return flat_cost, f"{product_code} / {reason}"

    return None, f"{product_code} matched but no price found at all"
