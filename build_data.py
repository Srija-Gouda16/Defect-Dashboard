"""
Parses the DEFECT LOG sheet from each downloaded xlsx file and builds the
combined dataset, summary stats, and insights JSON used by the dashboard
template. This is pure data-processing logic with no network calls, so it
can be tested independently of the OneDrive download step.
"""
import collections
import difflib
import json
import os
import re
import time
import zipfile
from datetime import datetime, date, timedelta

import openpyxl


# Formats we've seen or might reasonably see for a manually-typed date cell,
# tried in order. Excel usually gives us a real datetime/date object (handled
# separately below), but if someone types a date as plain text instead of
# picking it from a date picker, it comes through as a string in whatever
# format they happened to type - this is what silently dropped today's rows
# before, since only strict ISO format was accepted.
_DATE_STRING_FORMATS = [
    "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
    "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y", "%B %d, %Y", "%b %d, %Y",
]


def mask_sn(sn):
    """
    Masks a serial number for display: everything before the first dash
    becomes X's (same length, so the format still looks familiar), the
    dash and everything after it stays exactly as-is.
    e.g. '542633026832800-02178 0105 14568' -> 'XXXXXXXXXXXXXXX-02178 0105 14568'
    """
    if not sn:
        return sn
    sn = str(sn)
    if "-" not in sn:
        return "X" * len(sn)
    prefix, rest = sn.split("-", 1)
    return ("X" * len(prefix)) + "-" + rest


def normalize_date(value):
    """
    Returns a 'YYYY-MM-DD' string for any reasonable date representation
    Excel might hand us, or None if it genuinely isn't a date.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        # Excel serial date number (epoch 1899-12-30), in case formatting
        # broke and openpyxl handed us the raw number instead of a date.
        try:
            return (datetime(1899, 12, 30) + timedelta(days=value)).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            return None
    s = str(value).strip()
    if not s:
        return None
    # Already ISO, possibly with a time component attached
    if len(s) >= 10 and s[:4].isdigit() and s[4] == '-':
        return s[:10]
    for fmt in _DATE_STRING_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def clean_status(s):
    if not s:
        return "Unknown"
    s = str(s).strip()
    m = {
        "scrap": "Scrap", "re work": "Re work", "retest": "Retest",
        "ready for production": "Ready for production", "debug": "Debug",
        "reprocess": "Reprocess", "hold": "Hold", "sample": "Sample",
        "rework": "Re work", "re_work": "Re work",
    }
    return m.get(s.lower(), s)


def clean_shift(s):
    if not s:
        return "Unknown"
    s = str(s).strip().upper()
    return s if s in ("A", "B", "C") else "Unknown"


def nice_label(line_name):
    label = line_name.replace("LINE-", "Line ").replace("LINE ", "Line ")
    return label.replace("BAT 1& BAT 2", "BAT 1 & 2")


def sort_key(line_name):
    m = re.search(r"(\d+)", line_name)
    if "BAT" in line_name:
        return (1, 0)
    return (0, int(m.group(1)) if m else 99)


def load_model_costs_from_excel(filepath):
    """
    Reads model costs from a simple Excel file - two columns, MODEL and
    COST (header names, case-insensitive), one row per model. Duplicate
    models are averaged. Returns {} if the file doesn't exist yet, so this
    is safe to call even before anyone's created it.
    """
    if not filepath or not os.path.exists(filepath):
        return {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(h).strip().upper() if h else "" for h in header]
    col_idx = {}
    for i, h in enumerate(header):
        if h and h not in col_idx:
            col_idx[h] = i

    model_col = col_idx.get("MODEL")
    cost_col = col_idx.get("COST")
    if model_col is None or cost_col is None:
        wb.close()
        raise ValueError(f"Cost file must have MODEL and COST columns - found: {header}")

    values = collections.defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[model_col]
        cost = row[cost_col]
        if not model or cost is None:
            continue
        model = str(model).strip()
        try:
            cost = float(str(cost).replace("$", "").replace(",", "").strip())
        except ValueError:
            continue
        values[model].append(cost)
    wb.close()

    return {m: round(sum(c) / len(c), 2) for m, c in values.items()}


def extract_defect_rows(filepath, line_name, retries=3, retry_delay=3):
    """
    Read the DEFECT LOG sheet from one workbook into a list of dicts.

    Retries a few times if the file can't be opened yet - this happens if
    OneDrive is mid-sync or Excel briefly holds an exclusive lock right after
    a save, and without a retry that would silently produce stale/partial
    data for that run instead of failing loudly or just waiting it out.
    """
    last_error = None
    for attempt in range(retries):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            break
        except (PermissionError, zipfile.BadZipFile) as e:
            last_error = e
            if attempt < retries - 1:
                time.sleep(retry_delay)
    else:
        raise RuntimeError(f"Could not open {filepath} after {retries} attempts "
                           f"(file may be locked or still syncing): {last_error}")

    ws = wb["DEFECT LOG"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(h).strip() if h else "" for h in header]
    # Use the FIRST occurrence of each header name, not the last - files can
    # have extra trailing/hidden columns (some of these sheets have 40+
    # total columns despite only ~14 being visibly used), and if a header
    # name happens to repeat anywhere in that range, always overwriting
    # with the last match can silently point at the wrong (often blank)
    # column instead of the real one.
    col_idx = {}
    for i, h in enumerate(header):
        if h and h not in col_idx:
            col_idx[h] = i

    rows = []
    incomplete_count = 0
    empty_streak = 0
    # Safety cutoff, but generous - the earlier 2000-row limit was too low and
    # caused real entries to be silently dropped on files with a large blank
    # gap before them. 20000 is high enough to comfortably cover realistic
    # gaps while still bounding worst-case runtime on pathological files.
    MAX_EMPTY_STREAK = 20000
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[col_idx["DATE"]] if "DATE" in col_idx else None
        date_str = normalize_date(raw_date)
        if date_str is None:
            empty_streak += 1
            if empty_streak > MAX_EMPTY_STREAK:
                break
            continue
        empty_streak = 0

        raw_sn = str(row[col_idx["SN"]]).strip() if "SN" in col_idx and row[col_idx["SN"]] else None
        raw_station = str(row[col_idx["STATION"]]).strip() if "STATION" in col_idx and row[col_idx["STATION"]] else None
        raw_model = row[col_idx["MODEL"]] if "MODEL" in col_idx else None
        raw_problem = row[col_idx["PROBLEM DESCRIPTION"]] if "PROBLEM DESCRIPTION" in col_idx else None

        # Skip rows that only have DATE/SHIFT filled in (e.g. someone started
        # typing a new entry but hasn't finished it yet) - not a real defect
        # record until at least one of these is actually filled in.
        if not (raw_sn or raw_station or raw_model or raw_problem):
            incomplete_count += 1
            continue

        rows.append({
            "LINE_FILE": line_name,
            "DATE": date_str,
            "SHIFT": clean_shift(row[col_idx.get("SHIFT")] if "SHIFT" in col_idx else None),
            "SN": raw_sn,
            "STATION": raw_station or "Unknown",
            "MODEL": raw_model,
            "UNIT STATUS": clean_status(row[col_idx["UNIT STATUS"]] if "UNIT STATUS" in col_idx else None),
            "PROBLEM DESCRIPTION": raw_problem,
            "LOCATION": (str(row[col_idx["LOCATION"]]).strip()
                         if "LOCATION" in col_idx and row[col_idx["LOCATION"]] else "Unknown"),
            # Only present in some files (e.g. LINE-4, BAT 1&2) - blank/None
            # for files without this column.
            "SUB ASSEMBLY": (str(row[col_idx["SUB ASSEMBLY"]]).strip()
                              if "SUB ASSEMBLY" in col_idx and row[col_idx["SUB ASSEMBLY"]] else None),
        })
    wb.close()
    return rows, incomplete_count


def extract_component_scrap_rows(filepath, line_name, retries=3, retry_delay=3):
    """
    Read the COMPONENT SCRAP sheet from one workbook into a list of dicts.
    This is a SEPARATE sheet from DEFECT LOG, with its own structure - one
    row per component-scrap event, not one row per unit. Not every DMT LOG
    file may have this sheet, so a missing sheet returns an empty list
    rather than raising, and the caller can just carry on with other lines.
    """
    last_error = None
    for attempt in range(retries):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            break
        except (PermissionError, zipfile.BadZipFile) as e:
            last_error = e
            if attempt < retries - 1:
                time.sleep(retry_delay)
    else:
        raise RuntimeError(f"Could not open {filepath} after {retries} attempts "
                           f"(file may be locked or still syncing): {last_error}")

    if "COMPONENT SCRAP" not in wb.sheetnames:
        wb.close()
        return [], 0

    ws = wb["COMPONENT SCRAP"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(h).strip() if h else "" for h in header]
    col_idx = {}
    for i, h in enumerate(header):
        if h and h not in col_idx:
            col_idx[h] = i

    rows = []
    incomplete_count = 0
    empty_streak = 0
    MAX_EMPTY_STREAK = 20000
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[col_idx["DATE"]] if "DATE" in col_idx else None
        date_str = normalize_date(raw_date)
        raw_part = str(row[col_idx["PART NUMBER"]]).strip() if "PART NUMBER" in col_idx and row[col_idx["PART NUMBER"]] else None
        raw_qty = row[col_idx["QUANTITY"]] if "QUANTITY" in col_idx else None

        if date_str is None and raw_part is None:
            empty_streak += 1
            if empty_streak > MAX_EMPTY_STREAK:
                break
            continue
        empty_streak = 0

        # A row needs a date, a part number, AND a numeric quantity to be a
        # usable record - a handful of rows in real files have just a
        # REASON FOR REJECTION filled in with everything else blank, which
        # isn't a completed entry.
        try:
            qty = float(raw_qty) if raw_qty is not None else None
        except (TypeError, ValueError):
            qty = None
        if date_str is None or raw_part is None or qty is None:
            incomplete_count += 1
            continue

        raw_part_no_rev = (str(row[col_idx["PART NUMBER WITHOUT REVISION"]]).strip()
                            if "PART NUMBER WITHOUT REVISION" in col_idx and row[col_idx["PART NUMBER WITHOUT REVISION"]]
                            else raw_part)

        rows.append({
            "LINE_FILE": line_name,
            "DATE_ONLY": date_str,
            "SHIFT": clean_shift(row[col_idx.get("SHIFT")] if "SHIFT" in col_idx else None),
            "QUANTITY": qty,
            "PART_NUMBER": raw_part,
            "PART_NUMBER_NO_REV": raw_part_no_rev,
            "REASON": (str(row[col_idx["REASON FOR REJECTION"]]).strip()
                       if "REASON FOR REJECTION" in col_idx and row[col_idx["REASON FOR REJECTION"]] else None),
            "BOX_NUMBER": (str(row[col_idx["BOX NUMBER"]]).strip()
                           if "BOX NUMBER" in col_idx and row[col_idx["BOX NUMBER"]] else None),
        })
    wb.close()
    return rows, incomplete_count


def load_component_costs(filepath):
    """
    Weighted-average MFG_TX1 unit cost per Salcomp Part Number, from the
    master inventory layer cost report - same methodology as BOM costing:
    sum(layer value) / sum(layer quantity), across ALL lots for that org,
    not just the latest lot. Returns {} if the file can't be found/opened
    (component costs are then simply left blank rather than crashing the
    whole pipeline run over one missing/locked cost file).
    """
    if not filepath or not os.path.exists(filepath):
        return {}
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except (PermissionError, zipfile.BadZipFile):
        return {}

    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    agg_qty = collections.defaultdict(float)
    agg_val = collections.defaultdict(float)
    header_row = 7  # matches the known layout: real header on row 8 (1-indexed), data from row 8 on
    for row in ws.iter_rows(min_row=8, values_only=True):
        if len(row) < 14:
            continue
        org, salcode, qty, val = row[0], row[7], row[12], row[13]
        if org != "MFG_TX1" or not salcode:
            continue
        try:
            qty = float(qty)
            val = float(val)
        except (TypeError, ValueError):
            continue
        agg_qty[salcode] += qty
        agg_val[salcode] += val
    wb.close()

    return {part: (agg_val[part] / agg_qty[part]) for part in agg_qty if agg_qty[part] > 0}


def build_all(file_paths, model_costs=None, cost_trend_start_week=None, stage_cost_data=None,
              component_costs=None, component_aliases=None):
    """
    file_paths: dict of {line_name: local_filepath_or_BytesIO}, e.g.
        {"LINE-1": "/tmp/DMT LOG LINE-1.xlsx", "LINE 6": "/tmp/DMT LOG LINE 6.xlsx", ...}
    model_costs: optional dict of {model_name: unit_cost} - flat fallback cost,
        used when stage_cost_data doesn't have a match for a given row.
    cost_trend_start_week: optional string like "2026-W27" - if given, the
        Scrap Cost weekly trend chart only shows this week onward.
    stage_cost_data: optional dict from stage_cost.load_stage_cost_data() -
        if provided, scrap cost is calculated per the SPECIFIC stage a unit
        reached before failing (station-aware), not a flat per-model cost.
    component_costs: optional dict from load_component_costs() - weighted
        MFG_TX1 unit cost per Salcomp Part Number, used to cost the
        Component Scrap sheet's part-level scrap events.
    component_aliases: optional dict of {typo_as_in_dmt_log: correct_part_number}
        - explicit manual corrections (see config.COMPONENT_PART_ALIASES),
        checked before any automatic matching.

    Returns (dashboard_data, insights_data) ready to json.dumps into the template.
    """
    all_rows = []
    total_incomplete = 0
    component_scrap_rows = []
    total_component_incomplete = 0
    for line_name, path in file_paths.items():
        rows, incomplete = extract_defect_rows(path, line_name)
        all_rows.extend(rows)
        total_incomplete += incomplete
        comp_rows, comp_incomplete = extract_component_scrap_rows(path, line_name)
        component_scrap_rows.extend(comp_rows)
        total_component_incomplete += comp_incomplete
    return build_from_rows(all_rows, model_costs=model_costs, cost_trend_start_week=cost_trend_start_week,
                            incomplete_rows_dropped=total_incomplete, stage_cost_data=stage_cost_data,
                            component_scrap_rows=component_scrap_rows, component_costs=component_costs,
                            component_scrap_incomplete=total_component_incomplete,
                            component_aliases=component_aliases)


def build_from_rows(all_rows, model_costs=None, cost_trend_start_week=None, incomplete_rows_dropped=0,
                     stage_cost_data=None, component_scrap_rows=None, component_costs=None,
                     component_scrap_incomplete=0, component_aliases=None):
    """
    Same as build_all, but takes an already-extracted flat list of row
    dicts instead of file paths - used when rows came from somewhere other
    than the xlsx parser (e.g. csv_ingest.py reading a Power Automate CSV
    export). Each dict must have the same shape extract_defect_rows
    produces: LINE_FILE, DATE, SHIFT, SN, STATION, MODEL, UNIT STATUS,
    PROBLEM DESCRIPTION.
    """
    model_costs = model_costs or {}
    import stage_cost as _stage_cost_module

    def get_cost(model):
        if not model:
            return None
        model_str = str(model).strip()
        # Try the full string first
        if model_str in model_costs:
            return model_costs[model_str]
        # Fall back to trying each comma-separated segment - handles both
        # "E/M Assembly, IQ8BL-DOM" (real model after the comma) and
        # "IQ9N-A-INT, MADE IN US (INTERNAL USE ONLY)" (real model before
        # the comma), without needing to hardcode either pattern.
        if "," in model_str:
            for segment in model_str.split(","):
                segment = segment.strip()
                if segment in model_costs:
                    return model_costs[segment]
        return None

    def get_row_cost(model, station):
        """
        The cost for one specific scrap event. Tries stage-based costing
        first (station-aware - a unit failing early costs less than one
        failing late), falling back to the flat per-model cost file if
        stage costing isn't available or doesn't have a match.
        """
        if stage_cost_data:
            cost, _ = _stage_cost_module.get_scrap_cost(model, station, stage_cost_data)
            if cost is not None:
                return cost
        return get_cost(model)

    def get_row_stage(model, station):
        """Which build stage (SMT/MI/PCBA/Assembly Phantom/M-Assembly/
        Packaging/Cap/Adhesive/FG) a scrap event's station belongs to, for
        the wave-wise cost breakdown. Returns None if it can't be
        determined (unmatched model, or station not in the classification)."""
        if not stage_cost_data:
            return None
        product_code = _stage_cost_module.get_product_code(model, stage_cost_data)
        if product_code is None:
            return None
        family = stage_cost_data["product_code_to_family"].get(product_code)
        return _stage_cost_module.get_stage_for_station(family, station)

    # DATE is already normalized to 'YYYY-MM-DD' (or the row was dropped) by
    # extract_defect_rows, so no re-filtering needed here.
    good = all_rows
    for d in good:
        d["DATE_ONLY"] = d["DATE"][:10]
        d["PROBLEM DESCRIPTION"] = (str(d["PROBLEM DESCRIPTION"]).strip().upper()
                                     if d["PROBLEM DESCRIPTION"] else "UNSPECIFIED")

    max_date = max(d["DATE_ONLY"] for d in good)
    max_dt = datetime.strptime(max_date, "%Y-%m-%d")
    min_date = min(d["DATE_ONLY"] for d in good)

    # Daily view: rolling window of the last DAILY_WINDOW_DAYS days (not just
    # today) so there's always a meaningful amount of data to look at.
    DAILY_WINDOW_DAYS = 14
    daily_window_start = (max_dt - timedelta(days=DAILY_WINDOW_DAYS - 1)).strftime("%Y-%m-%d")

    lines_raw = sorted(set(d["LINE_FILE"] for d in good), key=sort_key)
    lines_order = ["ALL"] + lines_raw

    def top_problems_by_station(rows, top_problems=5, top_stations=5):
        """
        Top 5 problem descriptions, each broken down by which stations they
        occurred at - answers 'what's going wrong AND where' in one chart,
        instead of two separate ones.
        """
        problem_counts = collections.Counter(d["PROBLEM DESCRIPTION"] for d in rows if d["SN"])
        top_probs = [p for p, _ in problem_counts.most_common(top_problems)]
        sub_rows = [d for d in rows if d["PROBLEM DESCRIPTION"] in top_probs]

        pair_counts = collections.Counter((d["PROBLEM DESCRIPTION"], d["STATION"]) for d in sub_rows)
        station_totals = collections.Counter(d["STATION"] for d in sub_rows)
        top_stations_list = [s for s, _ in station_totals.most_common(top_stations)]

        datasets = {}
        for s in top_stations_list:
            datasets[s] = [pair_counts.get((p, s), 0) for p in top_probs]

        other_counts = []
        for p in top_probs:
            p_total = problem_counts[p]
            top_stat_total = sum(pair_counts.get((p, s), 0) for s in top_stations_list)
            other_counts.append(p_total - top_stat_total)
        if any(c > 0 for c in other_counts):
            datasets["Other"] = other_counts

        return {"problems": top_probs, "stations": list(datasets.keys()), "datasets": datasets}

    def top_locations_by_problem(rows, top_locations=5, top_problems=5):
        """
        Top 5 LOCATIONS (a more granular field than Station - e.g. 'Bulkhead',
        'PCBA', 'Barcode', 'Unit'), each broken down by problem description.
        """
        location_counts = collections.Counter(d["LOCATION"] for d in rows if d["SN"])
        top_locs = [l for l, _ in location_counts.most_common(top_locations)]
        sub_rows = [d for d in rows if d["LOCATION"] in top_locs]

        pair_counts = collections.Counter((d["LOCATION"], d["PROBLEM DESCRIPTION"]) for d in sub_rows)
        problem_totals = collections.Counter(d["PROBLEM DESCRIPTION"] for d in sub_rows)
        top_probs_list = [p for p, _ in problem_totals.most_common(top_problems)]

        datasets = {}
        for p in top_probs_list:
            datasets[p] = [pair_counts.get((l, p), 0) for l in top_locs]

        other_counts = []
        for l in top_locs:
            l_total = location_counts[l]
            top_prob_total = sum(pair_counts.get((l, p), 0) for p in top_probs_list)
            other_counts.append(l_total - top_prob_total)
        if any(c > 0 for c in other_counts):
            datasets["Other"] = other_counts

        return {"locations": top_locs, "problems": list(datasets.keys()), "datasets": datasets}

    def top_station_location_by_problem(rows, top_combos=5, top_problems=5):
        """
        All three fields in one chart: combines Station + Location into a
        single category (e.g. 'PCB VI / Bulkhead'), then breaks each down
        by problem description - where, the specific spot, and what went
        wrong, all at once.
        """
        def combo_label(d):
            station = d["STATION"] or "Unknown"
            location = d["LOCATION"] or "Unknown"
            return f"{station} / {location}"

        combo_counts = collections.Counter(combo_label(d) for d in rows if d["SN"])
        top_combos_list = [c for c, _ in combo_counts.most_common(top_combos)]
        sub_rows = [d for d in rows if combo_label(d) in top_combos_list]

        pair_counts = collections.Counter((combo_label(d), d["PROBLEM DESCRIPTION"]) for d in sub_rows)
        problem_totals = collections.Counter(d["PROBLEM DESCRIPTION"] for d in sub_rows)
        top_probs_list = [p for p, _ in problem_totals.most_common(top_problems)]

        datasets = {}
        for p in top_probs_list:
            datasets[p] = [pair_counts.get((c, p), 0) for c in top_combos_list]

        return {"combos": top_combos_list, "problems": list(datasets.keys()), "datasets": datasets}

    def build_summary(rows, trend_mode=None, trend_days=None):
        stations = collections.Counter(d["STATION"] for d in rows).most_common(8)
        shifts = collections.Counter(d["SHIFT"] for d in rows).most_common()
        status = collections.Counter(d["UNIT STATUS"] for d in rows).most_common()
        scrap = next((c for s, c in status if s == "Scrap"), 0)
        rework = next((c for s, c in status if s == "Re work"), 0)
        ready = next((c for s, c in status if s == "Ready for production"), 0)
        top_problems = top_problems_by_station(rows)
        top_locations = top_locations_by_problem(rows)
        top_combo = top_station_location_by_problem(rows)
        trend = []
        status_trend = {"scrap": [], "rework": [], "ready": []}
        if trend_mode == "daily":
            dc = collections.Counter(d["DATE_ONLY"] for d in rows)
            scrap_dc = collections.Counter(d["DATE_ONLY"] for d in rows if d["UNIT STATUS"] == "Scrap")
            rework_dc = collections.Counter(d["DATE_ONLY"] for d in rows if d["UNIT STATUS"] == "Re work")
            ready_dc = collections.Counter(d["DATE_ONLY"] for d in rows if d["UNIT STATUS"] == "Ready for production")
            for i in range(trend_days - 1, -1, -1):
                dt = (max_dt - timedelta(days=i)).strftime("%Y-%m-%d")
                trend.append([dt, dc.get(dt, 0)])
                status_trend["scrap"].append(scrap_dc.get(dt, 0))
                status_trend["rework"].append(rework_dc.get(dt, 0))
                status_trend["ready"].append(ready_dc.get(dt, 0))
        elif trend_mode == "weekly":
            # Bucket every record by ISO week, from the file's earliest date
            # through today - this is a real historical trend, not just a
            # fixed recent window.
            wc = collections.Counter()
            scrap_wc, rework_wc, ready_wc = collections.Counter(), collections.Counter(), collections.Counter()
            for d in rows:
                dt = datetime.strptime(d["DATE_ONLY"], "%Y-%m-%d")
                yr, wk, _ = dt.isocalendar()
                wkey = f"{yr}-W{wk:02d}"
                wc[wkey] += 1
                if d["UNIT STATUS"] == "Scrap": scrap_wc[wkey] += 1
                elif d["UNIT STATUS"] == "Re work": rework_wc[wkey] += 1
                elif d["UNIT STATUS"] == "Ready for production": ready_wc[wkey] += 1
            trend = sorted(wc.items())
            status_trend["scrap"] = [scrap_wc.get(w, 0) for w, _ in trend]
            status_trend["rework"] = [rework_wc.get(w, 0) for w, _ in trend]
            status_trend["ready"] = [ready_wc.get(w, 0) for w, _ in trend]
        return {"total": len(rows), "stations": stations, "shifts": shifts,
                "status": status, "scrap": scrap, "rework": rework, "ready": ready,
                "trend": trend, "status_trend": status_trend, "top_problems": top_problems, "top_locations": top_locations,
                "top_combo": top_combo}

    def table_rows(rows, limit=3000):
        rows_sorted = sorted(rows, key=lambda d: d["DATE_ONLY"], reverse=True)[:limit]
        return [{
            "date": d["DATE_ONLY"], "shift": d["SHIFT"], "sn": mask_sn(d["SN"]), "sn_full": d["SN"], "model": d["MODEL"],
            "station": d["STATION"], "status": d["UNIT STATUS"], "problem": d["PROBLEM DESCRIPTION"],
            "line": d["LINE_FILE"], "sub_assembly": d.get("SUB ASSEMBLY"), "location": d.get("LOCATION"),
        } for d in rows_sorted]

    def build_cost_view(rows, testers_scope_label):
        """
        Separate dedicated Scrap Cost view (not embedded in Hourly/Daily/
        Weekly): weekly scrap $ trend across full history, plus top 5
        highest-cost products by total scrap dollar impact.

        Cost is calculated PER ROW (not per model) since stage-based
        costing means the same model can cost different amounts depending
        on which station it failed at.
        """
        scrap_rows = [d for d in rows if d["UNIT STATUS"] == "Scrap" and d["MODEL"]]

        # Attach each scrap row's individual cost and build stage once, reused below
        for d in scrap_rows:
            d["_row_cost"] = get_row_cost(d["MODEL"], d["STATION"])
            d["_row_stage"] = get_row_stage(d["MODEL"], d["STATION"])

        # Component-level scrap (from the COMPONENT SCRAP sheet - raw parts
        # rejected before ever being built into a serialized unit) is a
        # separate, additive cost source from unit-level scrap above, since
        # a component scrap event has no SN and never overlaps with a
        # scrapped finished unit. Folded into every total below.
        line_component_rows = (component_scrap_rows if testers_scope_label == "ALL"
                                else [d for d in component_scrap_rows if d["LINE_FILE"] == testers_scope_label])

        # Weekly scrap $ trend, full history
        weekly_cost = collections.defaultdict(float)
        for d in scrap_rows:
            if d["_row_cost"] is None:
                continue
            dt = datetime.strptime(d["DATE_ONLY"], "%Y-%m-%d")
            yr, wk, _ = dt.isocalendar()
            weekly_cost[f"{yr}-W{wk:02d}"] += d["_row_cost"]
        for d in line_component_rows:
            dt = datetime.strptime(d["DATE_ONLY"], "%Y-%m-%d")
            yr, wk, _ = dt.isocalendar()
            weekly_cost[f"{yr}-W{wk:02d}"] += d["_ext_cost"]
        weekly_trend = sorted([[w, round(c, 2)] for w, c in weekly_cost.items()])
        if cost_trend_start_week:
            weekly_trend = [w for w in weekly_trend if w[0] >= cost_trend_start_week]

        # Daily scrap $ trend - last 2 days (yesterday + today), day by day
        daily_cost = collections.defaultdict(float)
        for d in scrap_rows:
            if d["_row_cost"] is not None:
                daily_cost[d["DATE_ONLY"]] += d["_row_cost"]
        for d in line_component_rows:
            daily_cost[d["DATE_ONLY"]] += d["_ext_cost"]
        daily_trend = []
        for i in range(1, -1, -1):
            dt = (max_dt - timedelta(days=i)).strftime("%Y-%m-%d")
            daily_trend.append([dt, round(daily_cost.get(dt, 0), 2)])

        # Separate 7-day daily series kept ONLY so the "Weekly" granularity
        # button's rolling "Last 7 days total" box still has enough days to
        # sum - the Daily chart/box above intentionally only shows 2 days.
        daily_trend_7 = []
        for i in range(6, -1, -1):
            dt = (max_dt - timedelta(days=i)).strftime("%Y-%m-%d")
            daily_trend_7.append([dt, round(daily_cost.get(dt, 0), 2)])

        # Monthly scrap $ trend - full history, bucketed by calendar month
        monthly_cost = collections.defaultdict(float)
        for d in scrap_rows:
            if d["_row_cost"] is not None:
                monthly_cost[d["DATE_ONLY"][:7]] += d["_row_cost"]  # "YYYY-MM"
        for d in line_component_rows:
            monthly_cost[d["DATE_ONLY"][:7]] += d["_ext_cost"]
        monthly_trend = sorted([[m, round(c, 2)] for m, c in monthly_cost.items()])

        # Top 5 highest-cost products, top 5 highest-cost products,
        # scrap cost by build stage - all follow the SAME Daily/Weekly/
        # Monthly granularity toggle as the trend chart above, so switching
        # the toggle updates all three together. Each sums each row's own
        # (possibly stage-specific) cost rather than unit_cost x count,
        # since two units of the same model can have different costs if
        # they failed at different stages.
        STAGE_ORDER = ["SMT", "MI", "PCBA", "Assembly Phantom", "M-Assembly",
                       "Packaging", "Cap", "Adhesive", "FG"]

        def compute_top5_and_stage(rows_subset, top_n=5):
            model_totals = collections.defaultdict(lambda: {"units": 0, "total_cost": 0.0})
            for d in rows_subset:
                model_totals[d["MODEL"]]["units"] += 1
                model_totals[d["MODEL"]]["total_cost"] += d["_row_cost"]
            entries = []
            for model, agg in model_totals.items():
                entries.append({
                    "model": model, "scrap_units": agg["units"],
                    "unit_cost": round(agg["total_cost"] / agg["units"], 4),
                    "total_cost": round(agg["total_cost"], 2),
                })
            entries.sort(key=lambda e: -e["total_cost"])
            top = entries[:top_n]

            stage_totals = collections.defaultdict(lambda: {"units": 0, "total_cost": 0.0})
            for d in rows_subset:
                stage_key = d["_row_stage"] or "Unclassified"
                stage_totals[stage_key]["units"] += 1
                stage_totals[stage_key]["total_cost"] += d["_row_cost"]
            wave = []
            remaining = dict(stage_totals)
            for stage in STAGE_ORDER:
                if stage in remaining:
                    agg = remaining.pop(stage)
                    wave.append({"stage": stage, "scrap_units": agg["units"], "total_cost": round(agg["total_cost"], 2)})
            for stage, agg in remaining.items():
                wave.append({"stage": stage, "scrap_units": agg["units"], "total_cost": round(agg["total_cost"], 2)})
            return top, wave

        # Daily period = same 2-day window as the Daily trend chart above
        daily_period_start = (max_dt - timedelta(days=1)).strftime("%Y-%m-%d")
        daily_period_rows = [d for d in scrap_rows if d["_row_cost"] is not None
                              and daily_period_start <= d["DATE_ONLY"] <= max_date]

        # Weekly period = same rolling last-7-days window as the Weekly
        # granularity's "Last 7 days total" box above
        weekly_period_start = (max_dt - timedelta(days=6)).strftime("%Y-%m-%d")
        weekly_period_rows = [d for d in scrap_rows if d["_row_cost"] is not None
                               and weekly_period_start <= d["DATE_ONLY"] <= max_date]

        # Monthly period = same last FULLY COMPLETE calendar month as the
        # Monthly granularity box above (skip the current in-progress month)
        months_present = sorted(monthly_cost.keys())
        complete_month = None
        if months_present:
            complete_month = months_present[-1]
            if complete_month == max_date[:7] and len(months_present) > 1:
                complete_month = months_present[-2]
        monthly_period_rows = [d for d in scrap_rows if d["_row_cost"] is not None
                                and complete_month and d["DATE_ONLY"][:7] == complete_month]

        top5_daily, wave_daily = compute_top5_and_stage(daily_period_rows)
        top5_weekly, wave_weekly = compute_top5_and_stage(weekly_period_rows)
        top5_monthly, wave_monthly = compute_top5_and_stage(monthly_period_rows)

        top5 = {"daily": top5_daily, "weekly": top5_weekly, "monthly": top5_monthly}
        wave_wise_cost = {"daily": wave_daily, "weekly": wave_weekly, "monthly": wave_monthly}

        # Q3 total scrap cost (July, August, September - calendar quarter,
        # any year present in the data). This KPI box stays fixed regardless
        # of the Daily/Weekly/Monthly toggle - separate from the above.
        # Includes component scrap $ for the same quarter/line.
        q3_cost = sum(d["_row_cost"] for d in scrap_rows
                      if d["_row_cost"] is not None and int(d["DATE_ONLY"][5:7]) in (7, 8, 9))
        q3_component_cost = sum(d["_ext_cost"] for d in line_component_rows if int(d["DATE_ONLY"][5:7]) in (7, 8, 9))
        q3_unit_scrap_cost = round(q3_cost, 2)
        q3_component_scrap_cost = round(q3_component_cost, 2)
        q3_total_cost = round(q3_cost + q3_component_cost, 2)

        # Q2 total scrap cost (April, May, June)
        q2_cost = sum(d["_row_cost"] for d in scrap_rows
                      if d["_row_cost"] is not None and int(d["DATE_ONLY"][5:7]) in (4, 5, 6))
        q2_component_cost = sum(d["_ext_cost"] for d in line_component_rows if int(d["DATE_ONLY"][5:7]) in (4, 5, 6))
        q2_total_cost = round(q2_cost + q2_component_cost, 2)

        # "Unclassified" diagnostic breakdown stays scoped to Q3, same as
        # before - it's a separate diagnostic table, not part of the
        # Daily/Weekly/Monthly toggle.
        q3_scrap_rows = [d for d in scrap_rows if int(d["DATE_ONLY"][5:7]) in (7, 8, 9) and d["_row_cost"] is not None]

        # What's actually inside "Unclassified" - the exact Station+Model
        # combos that couldn't be matched to a stage, so it's obvious what
        # needs fixing (either the model needs a Product Code mapping, or
        # the station needs adding to the classification list) instead of
        # just seeing an unexplained dollar figure.
        unclassified_rows = [d for d in q3_scrap_rows if d["_row_stage"] is None]
        unclassified_totals = collections.defaultdict(lambda: {"units": 0, "total_cost": 0.0})
        for d in unclassified_rows:
            key = (d["STATION"] or "Unknown", d["MODEL"])
            unclassified_totals[key]["units"] += 1
            unclassified_totals[key]["total_cost"] += d["_row_cost"]
        unclassified_breakdown = []
        for (station, model), agg in unclassified_totals.items():
            unclassified_breakdown.append({
                "station": station, "model": model, "scrap_units": agg["units"],
                "total_cost": round(agg["total_cost"], 2),
            })
        unclassified_breakdown.sort(key=lambda e: -e["total_cost"])
        unclassified_breakdown = unclassified_breakdown[:15]

        # Combined Q3 total (unit + component scrap), consistent with the
        # rest of the dashboard's Q3 framing - no all-time totals needed.
        total_scrap_cost = q3_total_cost

        return {
            "total_scrap_cost": total_scrap_cost,
            "q3_unit_scrap_cost": q3_unit_scrap_cost,
            "q3_component_scrap_cost": q3_component_scrap_cost,
            "q3_total_cost": q3_total_cost,
            "q2_total_cost": q2_total_cost,
            "weekly_trend": weekly_trend,
            "daily_trend": daily_trend,
            "daily_trend_7": daily_trend_7,
            "monthly_trend": monthly_trend,
            "top5_models": top5,
            "wave_wise_cost": wave_wise_cost,
            "unclassified_breakdown": unclassified_breakdown,
        }

    # Hourly view: today + yesterday (rolling 2-day window), per explicit
    # request - e.g. on the 5th it shows 4th+5th, on the 6th it shows 5th+6th.
    hourly_window_start = (max_dt - timedelta(days=1)).strftime("%Y-%m-%d")
    hourly_window_rows = [d for d in good if hourly_window_start <= d["DATE_ONLY"] <= max_date]
    hourly = {}
    hourly["ALL"] = {"label": "All lines", **build_summary(hourly_window_rows)}
    for l in lines_raw:
        hourly[l] = {"label": nice_label(l), **build_summary([d for d in hourly_window_rows if d["LINE_FILE"] == l])}

    daily = {}
    daily_window_rows = [d for d in good if daily_window_start <= d["DATE_ONLY"] <= max_date]
    daily["ALL"] = {"label": "All lines", **build_summary(daily_window_rows, trend_mode="daily", trend_days=DAILY_WINDOW_DAYS)}
    for l in lines_raw:
        daily[l] = {"label": nice_label(l), **build_summary(
            [d for d in daily_window_rows if d["LINE_FILE"] == l], trend_mode="daily", trend_days=DAILY_WINDOW_DAYS)}

    # Monthly view: rolling 30-day window, sits between Daily (14 days) and
    # Weekly (all-time) - a wider recent-activity lens than Daily without
    # going all the way back to the start of the file.
    MONTHLY_WINDOW_DAYS = 30
    monthly_window_start = (max_dt - timedelta(days=MONTHLY_WINDOW_DAYS - 1)).strftime("%Y-%m-%d")
    monthly_window_rows = [d for d in good if monthly_window_start <= d["DATE_ONLY"] <= max_date]
    monthly = {}
    monthly["ALL"] = {"label": "All lines", **build_summary(monthly_window_rows, trend_mode="daily", trend_days=MONTHLY_WINDOW_DAYS)}
    for l in lines_raw:
        monthly[l] = {"label": nice_label(l), **build_summary(
            [d for d in monthly_window_rows if d["LINE_FILE"] == l], trend_mode="daily", trend_days=MONTHLY_WINDOW_DAYS)}

    # Weekly view: everything from the start of the file through today,
    # trended by calendar week rather than a fixed recent window.
    weekly = {}
    weekly["ALL"] = {"label": "All lines", **build_summary(good, trend_mode="weekly")}
    for l in lines_raw:
        weekly[l] = {"label": nice_label(l), **build_summary([d for d in good if d["LINE_FILE"] == l], trend_mode="weekly")}

    # Component Scrap view: top 5 scrapped components by $ cost impact, from
    # the dedicated COMPONENT SCRAP sheet in each DMT log file (a SEPARATE
    # sheet from DEFECT LOG, with its own quantity/part-number records - not
    # a per-unit-row structure). Cost = quantity x MFG_TX1 weighted-average
    # unit cost, matched by Salcomp Part Number, tried in this order:
    #   1. Exact PART NUMBER match
    #   2. "PART NUMBER WITHOUT REVISION" (strips trailing _01/_02/etc)
    #   3. Base-code match: everything up through "EP" only, ignoring
    #      whatever revision suffix follows - if multiple priced revisions
    #      share that same base (e.g. UA05768EP_03 vs the priced
    #      UA05768EP_10), their costs are averaged. Confirmed correct by
    #      Srija for parts like this where the DMT log's revision number
    #      doesn't line up with what's currently priced in inventory.
    # Parts with no match at any of these three tiers are still counted by
    # quantity but contribute $0 - flagged separately below rather than
    # silently treated as free.
    component_costs = component_costs or {}
    component_scrap_rows = component_scrap_rows or []
    component_aliases = component_aliases or {}
    uncosted_component_parts = set()

    def base_ep_key(part_number):
        """Everything up through (and including) the last 'EP' in the code -
        e.g. 'UA05768EP_03' -> 'UA05768EP'. None if 'EP' isn't present."""
        if not part_number:
            return None
        idx = part_number.rfind("EP")
        return part_number[:idx + 2] if idx != -1 else None

    base_ep_costs = collections.defaultdict(list)
    for part, cost in component_costs.items():
        key = base_ep_key(part)
        if key:
            base_ep_costs[key].append(cost)
    base_ep_avg_cost = {key: sum(costs) / len(costs) for key, costs in base_ep_costs.items()}

    def get_component_cost(part_number, part_number_no_rev):
        # 0. Explicit manual correction for a known DMT log typo, if any
        aliased = component_aliases.get(part_number) or component_aliases.get(part_number_no_rev)
        if aliased and aliased in component_costs:
            return component_costs[aliased]
        if part_number in component_costs:
            return component_costs[part_number]
        if part_number_no_rev in component_costs:
            return component_costs[part_number_no_rev]
        key = base_ep_key(part_number) or base_ep_key(part_number_no_rev)
        if key in base_ep_avg_cost:
            return base_ep_avg_cost[key]
        return None

    for d in component_scrap_rows:
        unit_cost = get_component_cost(d["PART_NUMBER"], d["PART_NUMBER_NO_REV"])
        if unit_cost is None:
            uncosted_component_parts.add(d["PART_NUMBER"])
            d["_ext_cost"] = 0.0
        else:
            d["_ext_cost"] = d["QUANTITY"] * unit_cost

    # For parts with no cost match, suggest a close spelling match against
    # known-priced part numbers - purely diagnostic (shown to Srija so she
    # can catch DMT log typos like "CNO2374" meaning "CN02374EP"). NEVER
    # used to actually calculate cost - a wrong guess here would silently
    # assign the wrong price to the wrong part, which is worse than $0.
    uncosted_component_suggestions = {}
    if uncosted_component_parts and component_costs:
        known_parts = list(component_costs.keys())
        for part in uncosted_component_parts:
            matches = difflib.get_close_matches(part, known_parts, n=1, cutoff=0.75)
            if matches:
                uncosted_component_suggestions[part] = matches[0]

    component_daily_rows = [d for d in component_scrap_rows if daily_window_start <= d["DATE_ONLY"] <= max_date]
    component_monthly_rows = [d for d in component_scrap_rows if monthly_window_start <= d["DATE_ONLY"] <= max_date]
    component_q3_rows = [d for d in component_scrap_rows if int(d["DATE_ONLY"][5:7]) in (7, 8, 9)]

    def top_components(rows, top_n=5, rank_by="cost"):
        qty_totals = collections.defaultdict(float)
        cost_totals = collections.defaultdict(float)
        for d in rows:
            qty_totals[d["PART_NUMBER"]] += d["QUANTITY"]
            cost_totals[d["PART_NUMBER"]] += d["_ext_cost"]
        totals = qty_totals if rank_by == "qty" else cost_totals
        ranked = sorted(totals.items(), key=lambda kv: -kv[1])[:top_n]
        return [{"part": part, "qty": round(qty_totals[part], 2), "cost": round(cost_totals[part], 2)} for part, _ in ranked]

    def q3_total_qty(rows, line=None):
        subset = rows if line is None else [d for d in rows if d["LINE_FILE"] == line]
        return round(sum(d["QUANTITY"] for d in subset), 2)

    component_view = {}
    component_view["ALL"] = {
        "label": "All lines",
        "daily_components": top_components(component_daily_rows),
        "weekly_components": top_components(component_scrap_rows),
        "monthly_components": top_components(component_monthly_rows),
        "daily_components_qty": top_components(component_daily_rows, rank_by="qty"),
        "weekly_components_qty": top_components(component_scrap_rows, rank_by="qty"),
        "monthly_components_qty": top_components(component_monthly_rows, rank_by="qty"),
        "q3_total_qty": q3_total_qty(component_q3_rows),
    }
    for l in lines_raw:
        component_view[l] = {
            "label": nice_label(l),
            "daily_components": top_components([d for d in component_daily_rows if d["LINE_FILE"] == l]),
            "weekly_components": top_components([d for d in component_scrap_rows if d["LINE_FILE"] == l]),
            "monthly_components": top_components([d for d in component_monthly_rows if d["LINE_FILE"] == l]),
            "daily_components_qty": top_components([d for d in component_daily_rows if d["LINE_FILE"] == l], rank_by="qty"),
            "weekly_components_qty": top_components([d for d in component_scrap_rows if d["LINE_FILE"] == l], rank_by="qty"),
            "monthly_components_qty": top_components([d for d in component_monthly_rows if d["LINE_FILE"] == l], rank_by="qty"),
            "q3_total_qty": q3_total_qty(component_q3_rows, line=l),
        }

    # Hourly/live feed: today + yesterday, most recent first.
    hourly_sorted = sorted(hourly_window_rows, key=lambda d: d["DATE_ONLY"], reverse=True)
    live_feed_table = table_rows(hourly_sorted, limit=2000)

    live_counts = {"ALL": {"label": "All lines", "total": len(hourly_window_rows)}}
    for l in lines_raw:
        live_counts[l] = {"label": nice_label(l), "total": len([d for d in hourly_window_rows if d["LINE_FILE"] == l])}

    # Filter dropdown options: computed from the FULL history for each line
    # (not limited to whatever window/table happens to be showing), so every
    # value that's ever appeared is selectable regardless of which view or
    # date range you're currently looking at.
    def filter_options_for(rows):
        return {
            "dates": sorted(set(d["DATE_ONLY"] for d in rows), reverse=True),
            "shifts": sorted(set(d["SHIFT"] for d in rows if d["SHIFT"])),
            "stations": sorted(set(d["STATION"] for d in rows if d["STATION"])),
            "models": sorted(set(d["MODEL"] for d in rows if d["MODEL"])),
            "statuses": sorted(set(d["UNIT STATUS"] for d in rows if d["UNIT STATUS"])),
        }

    filter_options = {"ALL": filter_options_for(good)}
    for l in lines_raw:
        filter_options[l] = filter_options_for([d for d in good if d["LINE_FILE"] == l])

    # Dedicated Scrap Cost view - separate from Hourly/Daily/Weekly, full
    # history, one entry per line + ALL.
    cost_view = {}
    if model_costs or stage_cost_data:
        cost_view["ALL"] = {"label": "All lines", **build_cost_view(good, "ALL")}
        for l in lines_raw:
            cost_view[l] = {"label": nice_label(l), **build_cost_view([d for d in good if d["LINE_FILE"] == l], l)}

    # Data quality: how many rows got dropped for having only date/shift
    # filled in - surfaces entry problems instead of silently hiding them.
    # Models that appear in scrapped rows but have no cost on file - lets
    # you know your cost list is missing something instead of just
    # silently showing a lower total than reality.
    uncosted_models = []
    if model_costs or stage_cost_data:
        scrap_rows_for_uncosted = [d for d in good if d["UNIT STATUS"] == "Scrap" and d["MODEL"]]
        scrap_model_counts = collections.Counter(d["MODEL"] for d in scrap_rows_for_uncosted)
        for model, count in scrap_model_counts.most_common():
            sample_row = next(d for d in scrap_rows_for_uncosted if d["MODEL"] == model)
            if get_row_cost(model, sample_row["STATION"]) is None:
                uncosted_models.append({"model": model, "scrap_units": count})
        uncosted_models = uncosted_models[:10]

    dashboard_data = {
        "lines_order": lines_order, "max_date": max_date, "min_date": min_date,
        "hourly_window_start": hourly_window_start,
        "daily_window_start": daily_window_start, "daily_window_days": DAILY_WINDOW_DAYS,
        "monthly_window_start": monthly_window_start, "monthly_window_days": MONTHLY_WINDOW_DAYS,
        "generated_at": datetime.now().strftime("%Y-%m-%d %I:%M:%S %p"),
        "hourly": hourly, "daily": daily, "monthly": monthly, "weekly": weekly, "live_counts": live_counts,
        "filter_options": filter_options,
        "live_feed_table": live_feed_table,
        "daily_table": table_rows(daily_window_rows),
        "monthly_table": table_rows(monthly_window_rows),
        "weekly_table": table_rows(good),
        "cost_view": cost_view,
        "component_view": component_view,
        "incomplete_rows_dropped": incomplete_rows_dropped,
        "uncosted_models": uncosted_models,
        "uncosted_component_parts": sorted(uncosted_component_parts),
        "uncosted_component_suggestions": uncosted_component_suggestions,
        "component_scrap_incomplete_dropped": component_scrap_incomplete,
        "component_scrap_total_rows": len(component_scrap_rows),
    }

    # ---- Insights: Pareto, repeat offenders, line/shift rates, trend alerts ----
    def pareto(rows, top_n=10):
        cnt = collections.Counter(d["PROBLEM DESCRIPTION"] for d in rows if d["SN"])
        total = sum(cnt.values())
        out, cum = [], 0
        for name, c in cnt.most_common(top_n):
            cum += c
            out.append([name, c, round(100 * c / total, 1) if total else 0,
                        round(100 * cum / total, 1) if total else 0])
        return out

    def repeat_offenders(rows, min_count=3, top_n=15):
        cnt = collections.Counter(d["SN"] for d in rows if d["SN"])
        offenders = sorted([(sn, c) for sn, c in cnt.items() if c >= min_count], key=lambda x: -x[1])
        result = []
        for sn, c in offenders[:top_n]:
            sub = [d for d in rows if d["SN"] == sn]
            result.append({
                "sn": mask_sn(sn), "count": c, "model": sub[0]["MODEL"],
                "stations": sorted(set(d["STATION"] for d in sub)),
                "last_status": sub[-1]["UNIT STATUS"],
            })
        return result

    pareto_all = {"ALL": pareto(good)}
    repeat_all = {"ALL": repeat_offenders(good)}
    for l in lines_raw:
        rows = [d for d in good if d["LINE_FILE"] == l]
        pareto_all[l] = pareto(rows)
        repeat_all[l] = repeat_offenders(rows)

    line_rates = {}
    for l in lines_raw:
        rows = [d for d in good if d["LINE_FILE"] == l]
        sns = set(d["SN"] for d in rows if d["SN"])
        scrap_sns = set(d["SN"] for d in rows if d["SN"] and d["UNIT STATUS"] == "Scrap")
        line_rates[l] = {
            "label": nice_label(l), "unique_units": len(sns), "scrap_units": len(scrap_sns),
            "scrap_rate": round(100 * len(scrap_sns) / len(sns), 1) if sns else 0,
        }

    def shift_rate(rows):
        out = {}
        for sh in ["A", "B", "C"]:
            sub = [d for d in rows if d["SHIFT"] == sh]
            scrap = len([d for d in sub if d["UNIT STATUS"] == "Scrap"])
            out[sh] = {"total": len(sub), "scrap": scrap,
                       "scrap_rate": round(100 * scrap / len(sub), 1) if sub else 0}
        return out

    shift_rates = {"all": shift_rate(good), "per_line": {l: shift_rate([d for d in good if d["LINE_FILE"] == l]) for l in lines_raw}}

    alerts_total, alerts_scrap = [], []
    for l in lines_raw:
        rows = [d for d in good if d["LINE_FILE"] == l]
        daily_counts = collections.Counter(d["DATE_ONLY"] for d in rows)
        daily_scrap = collections.Counter(d["DATE_ONLY"] for d in rows if d["UNIT STATUS"] == "Scrap")
        prior_days = [(max_dt - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(1, 8)]

        def make_alert(counter):
            today_val = counter.get(max_date, 0)
            prior_vals = [counter.get(pd, 0) for pd in prior_days]
            avg = sum(prior_vals) / len(prior_vals) if prior_vals else 0
            pct = round(100 * (today_val - avg) / avg, 1) if avg > 0 else (100 if today_val > 0 else 0)
            status = "high" if pct >= 30 else ("low" if pct <= -30 else "normal")
            return {"line": l, "label": nice_label(l), "today": today_val, "avg_7day": round(avg, 1),
                    "pct_change": pct, "status": status}

        alerts_total.append(make_alert(daily_counts))
        alerts_scrap.append(make_alert(daily_scrap))

    insights_data = {
        "pareto": pareto_all, "repeat_offenders": repeat_all,
        "line_rates": line_rates, "shift_rates": shift_rates,
        "alerts": {"total": alerts_total, "scrap": alerts_scrap},
    }

    return dashboard_data, insights_data
